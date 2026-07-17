"""
xlsx_parser.py — XLSX / XLS extraction using openpyxl.

Each worksheet becomes:
  - A Markdown table section
  - A JSON table with sheet name, headers, and rows

Preserves:
  - Sheet names
  - Column headers (first row treated as header)
  - Row numbers (1-indexed, 1 = header)
  - Empty cell handling
"""

from __future__ import annotations

import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from .logger import get_logger, log_event

log = get_logger()

# Max rows to process per sheet (safety cap for enormous govt spreadsheets)
_MAX_ROWS = 2000


# ─────────────────────────────────────────────────────────────────────────────
# Result dataclass
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class SheetResult:
    sheet_name: str = ""
    headers: list[str] = field(default_factory=list)
    rows: list[list[Any]] = field(default_factory=list)
    row_count: int = 0
    col_count: int = 0

    def to_markdown(self) -> str:
        """Render this sheet as a Markdown table."""
        if not self.headers:
            return f"### {self.sheet_name}\n\n*(empty sheet)*\n"

        lines: list[str] = [f"### Sheet: {self.sheet_name}\n"]
        # Header row
        lines.append("| " + " | ".join(str(h) for h in self.headers) + " |")
        # Separator
        lines.append("| " + " | ".join("---" for _ in self.headers) + " |")
        # Data rows
        for row in self.rows:
            cells = [str(c).replace("|", "\\|") if c is not None else "" for c in row]
            # Pad to header width
            while len(cells) < len(self.headers):
                cells.append("")
            lines.append("| " + " | ".join(cells) + " |")

        lines.append("")
        return "\n".join(lines)

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "headers":    self.headers,
            "row_count":  self.row_count,
            "col_count":  self.col_count,
            "rows":       self.rows,
        }


@dataclass
class XlsxParseResult:
    sheets: list[SheetResult] = field(default_factory=list)
    markdown: str = ""
    json_data: dict[str, Any] = field(default_factory=dict)
    table_count: int = 0
    word_count: int = 0
    parser_used: str = "openpyxl"
    duration_s: float = 0.0
    success: bool = False
    error: str = ""


class XlsxExtractionError(Exception):
    """Raised when XLSX extraction fails."""


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def parse_xlsx(
    xlsx_path: Path,
    *,
    max_rows: int = _MAX_ROWS,
    url: str = "",
) -> XlsxParseResult:
    """
    Extract all worksheets from an XLSX (or XLS) file.

    Parameters
    ----------
    xlsx_path:
        Path to the downloaded spreadsheet.
    max_rows:
        Safety cap on rows per sheet.
    url:
        Original URL (for logging only).
    """
    t0 = time.monotonic()
    log_event(log, "PARSE", f"Parsing XLSX: {xlsx_path.name}", url=url)

    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl is not installed. Run: pip install openpyxl") from exc

    try:
        # openpyxl cannot read .xls (legacy Excel 97-2003)
        # We handle .xls by trying xlrd as a fallback
        if xlsx_path.suffix.lower() == ".xls":
            return _parse_xls(xlsx_path, max_rows=max_rows, url=url, t0=t0)

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        sheets: list[SheetResult] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet = _extract_sheet(ws, sheet_name, max_rows)
            if sheet.row_count > 0 or sheet.col_count > 0:
                sheets.append(sheet)

        wb.close()

        if not sheets:
            raise XlsxExtractionError(f"No data found in {xlsx_path.name}")

        # ── Assemble markdown ──────────────────────────────────────────
        md_parts = [f"# {xlsx_path.stem}\n\n"]
        for sheet in sheets:
            md_parts.append(sheet.to_markdown())
        markdown = "\n".join(md_parts)

        # ── Assemble JSON ──────────────────────────────────────────────
        json_data: dict[str, Any] = {
            "file": xlsx_path.name,
            "sheet_count": len(sheets),
            "sheets": [s.to_dict() for s in sheets],
        }

        elapsed = time.monotonic() - t0
        word_count = len(markdown.split())

        log_event(
            log, "PARSE",
            f"XLSX parsed: {len(sheets)} sheets · {word_count:,} words in {elapsed:.1f}s",
            url=url, duration_s=round(elapsed, 3),
        )

        return XlsxParseResult(
            sheets=sheets,
            markdown=markdown,
            json_data=json_data,
            table_count=len(sheets),
            word_count=word_count,
            parser_used="openpyxl",
            duration_s=round(elapsed, 3),
            success=True,
        )

    except XlsxExtractionError:
        raise
    except Exception as exc:
        elapsed = time.monotonic() - t0
        msg = f"{type(exc).__name__}: {exc}"
        log_event(log, "FAIL", f"XLSX parse error: {msg}", level="ERROR", url=url)
        raise XlsxExtractionError(msg) from exc


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _extract_sheet(ws, sheet_name: str, max_rows: int) -> SheetResult:
    """Extract one openpyxl worksheet into a SheetResult."""
    rows_iter = ws.iter_rows(values_only=True)

    # First row → headers
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return SheetResult(sheet_name=sheet_name)

    headers = [str(h) if h is not None else f"Col{i+1}" for i, h in enumerate(raw_headers)]
    # Remove trailing empty headers
    while headers and headers[-1].startswith("Col"):
        last_idx = int(headers[-1][3:]) - 1
        if raw_headers[last_idx] is None:
            headers.pop()
        else:
            break

    if not headers:
        return SheetResult(sheet_name=sheet_name)

    rows: list[list[Any]] = []
    for i, row_values in enumerate(rows_iter):
        if i >= max_rows:
            break
        # Trim to header width, convert to strings
        cells = list(row_values[:len(headers)])
        # Skip entirely empty rows
        if all(c is None or str(c).strip() == "" for c in cells):
            continue
        rows.append(cells)

    return SheetResult(
        sheet_name=sheet_name,
        headers=headers,
        rows=rows,
        row_count=len(rows),
        col_count=len(headers),
    )


def _parse_xls(xlsx_path: Path, *, max_rows: int, url: str, t0: float) -> XlsxParseResult:
    """
    Fallback parser for legacy .xls files using xlrd.
    If xlrd is not installed, raises XlsxExtractionError.
    """
    try:
        import xlrd
    except ImportError:
        raise XlsxExtractionError(
            f"Cannot parse .xls: xlrd is not installed. "
            f"Run: pip install xlrd==1.2.0\nURL: {url}"
        )

    try:
        wb = xlrd.open_workbook(str(xlsx_path))
        sheets: list[SheetResult] = []

        for ws in wb.sheets():
            if ws.nrows == 0:
                continue
            headers = [str(ws.cell_value(0, c)) or f"Col{c+1}" for c in range(ws.ncols)]
            rows: list[list[Any]] = []
            for r in range(1, min(ws.nrows, max_rows + 1)):
                row = [ws.cell_value(r, c) for c in range(ws.ncols)]
                if any(str(v).strip() for v in row):
                    rows.append(row)

            sheets.append(SheetResult(
                sheet_name=ws.name,
                headers=headers,
                rows=rows,
                row_count=len(rows),
                col_count=ws.ncols,
            ))

        if not sheets:
            raise XlsxExtractionError(f"No data in {xlsx_path.name}")

        md_parts = [f"# {xlsx_path.stem}\n\n"]
        for sheet in sheets:
            md_parts.append(sheet.to_markdown())
        markdown = "\n".join(md_parts)

        json_data: dict[str, Any] = {
            "file": xlsx_path.name,
            "sheet_count": len(sheets),
            "sheets": [s.to_dict() for s in sheets],
        }

        elapsed = time.monotonic() - t0
        return XlsxParseResult(
            sheets=sheets,
            markdown=markdown,
            json_data=json_data,
            table_count=len(sheets),
            word_count=len(markdown.split()),
            parser_used="xlrd",
            duration_s=round(elapsed, 3),
            success=True,
        )

    except XlsxExtractionError:
        raise
    except Exception as exc:
        raise XlsxExtractionError(f"xlrd error: {exc}") from exc
