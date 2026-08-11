"""Discover, download, and extract citizen-facing official documents.

The West Tripura portal links many documents from CDN/S3WAAS hosts instead of
hosting the binary on the district domain. This module keeps page crawling on
westtripura.nic.in while treating trusted document hosts as downloadable assets.
"""
from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from urllib.parse import urljoin, urlparse, parse_qs

import httpx
from bs4 import BeautifulSoup

OUTPUT_DIR = Path("output")
ASSET_DIR = OUTPUT_DIR / "documents"
ASSET_MANIFEST = OUTPUT_DIR / "documents.jsonl"
SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".doc", ".xlsx", ".xls", ".csv", ".txt"}
TRUSTED_HOSTS = {
    "westtripura.nic.in",
    "cdn.s3waas.gov.in",
    "s3waas.gov.in",
    "drive.google.com",
}


def _extension(url: str, content_type: str = "") -> str:
    path = urlparse(url).path.lower()
    for ext in SUPPORTED_EXTENSIONS:
        if path.endswith(ext):
            return ext
    ctype = content_type.lower()
    if "pdf" in ctype:
        return ".pdf"
    if "word" in ctype or "officedocument.word" in ctype:
        return ".docx"
    if "spreadsheet" in ctype or "excel" in ctype:
        return ".xlsx"
    if "csv" in ctype:
        return ".csv"
    return ""


def _trusted(url: str) -> bool:
    host = (urlparse(url).hostname or "").lower()
    return host in TRUSTED_HOSTS or host.endswith(".s3waas.gov.in")


def _normalise_url(base_url: str, href: str) -> str:
    href = href.strip()
    if not href or href.startswith(("javascript:", "mailto:", "tel:")):
        return ""
    url = urljoin(base_url, href)
    parsed = urlparse(url)
    # Google Drive links need a separate resolver at download time.
    return parsed._replace(fragment="").geturl()


def discover_document_links(html: str, page_url: str) -> list[dict]:
    """Return document links found in a crawled page, including trusted CDN links."""
    soup = BeautifulSoup(html or "", "html.parser")
    found: dict[str, dict] = {}
    for anchor in soup.find_all("a", href=True):
        url = _normalise_url(page_url, anchor.get("href", ""))
        if not url or not _trusted(url):
            continue
        ext = _extension(url)
        text = " ".join(anchor.stripped_strings)
        if ext or "download" in text.lower() or "view" in text.lower():
            found[url] = {"url": url, "title": text or url.rsplit("/", 1)[-1], "source_url": page_url}
    return list(found.values())


def _google_drive_download_url(url: str) -> str:
    parsed = urlparse(url)
    if parsed.hostname != "drive.google.com":
        return url
    match = re.search(r"/file/d/([^/]+)", parsed.path)
    if match:
        return f"https://drive.usercontent.google.com/download?id={match.group(1)}&export=download&confirm=t"
    file_id = parse_qs(parsed.query).get("id", [None])[0]
    if file_id:
        return f"https://drive.usercontent.google.com/download?id={file_id}&export=download&confirm=t"
    return url


def _safe_name(title: str, url: str, ext: str) -> str:
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", title).strip("._-")[:120] or "document"
    digest = hashlib.sha256(url.encode()).hexdigest()[:12]
    if not stem.lower().endswith(ext):
        stem += ext
    return f"{stem.rsplit(ext, 1)[0]}_{digest}{ext}"


async def download_documents(links: list[dict], timeout: float = 45.0) -> list[dict]:
    """Download discovered assets once and return manifest records."""
    ASSET_DIR.mkdir(parents=True, exist_ok=True)
    existing: dict[str, dict] = {}
    if ASSET_MANIFEST.exists():
        for line in ASSET_MANIFEST.read_text(encoding="utf-8").splitlines():
            try:
                row = json.loads(line)
                existing[row["url"]] = row
            except Exception:
                continue

    records: list[dict] = []
    headers = {"User-Agent": "WestTripura-RAG-ingestor/1.0"}
    async with httpx.AsyncClient(follow_redirects=True, timeout=timeout, headers=headers) as client:
        for item in links:
            source_url = item["url"]
            if source_url in existing and Path(existing[source_url].get("local_path", "")).exists():
                records.append(existing[source_url])
                continue
            try:
                response = await client.get(_google_drive_download_url(source_url))
                response.raise_for_status()
                ctype = response.headers.get("content-type", "")
                ext = _extension(str(response.url), ctype) or _extension(source_url)
                if not ext:
                    continue
                filename = _safe_name(item.get("title", "document"), source_url, ext)
                local_path = ASSET_DIR / filename
                local_path.write_bytes(response.content)
                row = {
                    "url": source_url,
                    "resolved_url": str(response.url),
                    "local_path": str(local_path),
                    "title": item.get("title", filename),
                    "source_url": item.get("source_url", ""),
                    "extension": ext,
                    "bytes": len(response.content),
                }
                records.append(row)
                existing[source_url] = row
            except Exception as exc:
                records.append({
                    "url": source_url,
                    "title": item.get("title", source_url),
                    "source_url": item.get("source_url", ""),
                    "error": str(exc),
                })

    with ASSET_MANIFEST.open("w", encoding="utf-8") as f:
        for row in {r["url"]: r for r in [*existing.values(), *records]}.values():
            f.write(json.dumps(row, ensure_ascii=False) + "\n")
    return records


def extract_document_text(path: str | Path) -> str:
    """Extract searchable text from supported downloaded document formats."""
    path = Path(path)
    ext = path.suffix.lower()
    if ext == ".pdf":
        from pypdf import PdfReader
        pages = []
        for page in PdfReader(str(path)).pages:
            pages.append(page.extract_text() or "")
        return "\n\n".join(pages).strip()
    if ext == ".docx":
        from docx import Document
        doc = Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip()).strip()
    if ext in {".xlsx", ".xls"}:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        rows = []
        for ws in wb.worksheets:
            rows.append(f"## Sheet: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                values = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if values:
                    rows.append(" | ".join(values))
        return "\n".join(rows).strip()
    if ext == ".csv":
        return path.read_text(encoding="utf-8", errors="ignore").strip()
    if ext == ".txt":
        return path.read_text(encoding="utf-8", errors="ignore").strip()
    if ext == ".doc":
        raise ValueError("Legacy .doc requires LibreOffice/docling for extraction")
    return ""
